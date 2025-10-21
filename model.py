from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import Optional, List, Any, ClassVar
import os

try:
	from openpyxl import Workbook, load_workbook
	from openpyxl.worksheet.worksheet import Worksheet
except Exception:  # defer import errors until save is actually called
	Workbook = None  # type: ignore
	load_workbook = None  # type: ignore
	Worksheet = None  # type: ignore


class OperationType(Enum):
	CLICK = "click"
	SCROLL = "scroll"
	INPUT_TEXT = "inputText"


@dataclass
class ActionSpec:
	"""Represents one UI action step for browsing automation."""
	# Required fields first (no defaults)
	url: str
	htmlComponent: str
	operationType: OperationType

	# Optional fields with defaults after required ones
	Input: Optional[str] = None
	# Auto-incrementing primary key; assigned if not provided
	id: Optional[int] = field(default=None, repr=True)

	def to_row(self) -> List[Any]:
		return [self.id, self.url, self.htmlComponent, self.operationType.value, self.Input]


def save_action_to_excel(action: ActionSpec, file_path: str = "actions.xlsx", sheet_name: str = "Actions") -> None:
	"""Save the action to an Excel file, maintaining rows in ascending id order.

	If the file/sheet doesn't exist, it will be created with a header row.
	"""
	if Workbook is None or load_workbook is None:
		raise RuntimeError("openpyxl is required to save to Excel. Install with: pip install openpyxl")

	headers = ["id", "url", "htmlComponent", "operationType", "Input"]

	# Load or create workbook and sheet
	if os.path.exists(file_path):
		wb = load_workbook(file_path)
		ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
	else:
		wb = Workbook()
		ws = wb.active
		ws.title = sheet_name
		ws.append(headers)

	# Collect existing rows (skip header if present)
	rows: List[List[Any]] = []
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		if i == 0 and list(row) == headers:
			continue
		rows.append(list(row))

	# Add current action
	if action.id is None:
		action.id = len(rows) + 1
	rows.append(action.to_row())

	# Rewrite sheet: clear and write header + sorted rows
	wb.remove(ws)
	ws = wb.create_sheet(title=sheet_name)
	ws.append(headers)
	for r in rows:
		ws.append(r)

	# Ensure the created sheet is the first if it's a new workbook
	if wb.sheetnames[0] != sheet_name:
		wb.move_sheet(sheet_name, offset=-wb.sheetnames.index(sheet_name))

	wb.save(file_path)


def get_all_actions_from_excel(file_path: str = "actions.xlsx", sheet_name: str = "Actions") -> List[ActionSpec]:
	"""Read all actions from Excel file and convert them to ActionSpec objects.
	
	Returns a list of ActionSpec objects from the Excel file.
	"""
	if load_workbook is None:
		raise RuntimeError("openpyxl is required to read from Excel. Install with: pip install openpyxl")
	
	if not os.path.exists(file_path):
		raise FileNotFoundError(f"Excel file not found: {file_path}")
	
	wb = load_workbook(file_path)
	if sheet_name not in wb.sheetnames:
		raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
	
	ws = wb[sheet_name]
	actions = []
	
	# Skip header row and process data rows
	for i, row in enumerate(ws.iter_rows(values_only=True), 1):
		if i == 1:  # Skip header row
			continue
		
		if not any(row):  # Skip empty rows
			continue
		
		try:
			# Extract data from row (assuming order: id, url, htmlComponent, operationType, Input)
			action_id = row[0] if row[0] is not None else None
			url = row[1] if row[1] is not None else ""
			html_component = row[2] if row[2] is not None else ""
			operation_type_str = row[3] if row[3] is not None else ""
			input_value = row[4] if row[4] is not None else None
			
			# Convert operation type string to enum
			operation_type = OperationType(operation_type_str)
			
			# Create ActionSpec object
			action = ActionSpec(
				id=action_id,
				url=url,
				htmlComponent=html_component,
				operationType=operation_type,
				Input=input_value
			)
			actions.append(action)
			
		except (ValueError, IndexError) as e:
			print(f"Warning: Skipping invalid row {i}: {e}")
			continue
	
	return actions
